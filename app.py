import streamlit as st
import pandas as pd
import openpyxl
from datetime import date, time
import io
import zipfile

st.set_page_config(page_title="Bitácoras por Promotor", page_icon="📋", layout="centered")
st.title("📋 Generador de Bitácoras por Promotor")

# ─────────────────────────────────────────────────────────────────
# CONFIGURACIÓN FIJA
# ─────────────────────────────────────────────────────────────────
CONTRATO_PROMOTOR = {
    9890:   "MIGUEL ANGEL TEBAR PEDROZA",   # CAPITALES FACILITATION
    100320: "MIGUEL ANGEL TEBAR PEDROZA",   # INDUSTRIAS CH
    100321: "MIGUEL ANGEL TEBAR PEDROZA",   # GRUPO SIMEC
    104351: "MIGUEL ANGEL TEBAR PEDROZA",   # BANCO AZTECA (H2H)
    105433: "MIGUEL ANGEL TEBAR PEDROZA",   # SEGUROS AZTECA
    105862: "MIGUEL ANGEL TEBAR PEDROZA",   # COMPASS INVESTMENTS
    106043: "MIGUEL ANGEL TEBAR PEDROZA",
    104871: "JOSE LUIS ALCAINE",            # FONDO DE PROMOCION B
    105775: "JOSE LUIS ALCAINE",            # SKANDIA LIFE
    100844: "JOSE LUIS ALCAINE",            # HDI SEGUROS
    105434: "JOSE LUIS ALCAINE",
    105777: "JOSE LUIS ALCAINE",
    106044: "GERARDO PEREZ CRUZ",
}

OP_TO_NAME = {
    "CB1074134":  "GERARDO PEREZ CRUZ",
    "CB1059258":  "MIGUEL ANGEL TEBAR PEDROZA",
    "CBP1059258": "MIGUEL ANGEL TEBAR PEDROZA",
    "1059258":    "MIGUEL ANGEL TEBAR PEDROZA",
    "CLCB178007": "ALBERTO ALARCON GONZALEZ",
    "CB331177":   "CB331177",
    "H2H":        "H2H",
}

LAYOUT_MAP = {
    "Fecha de la instrucción \n-recepción-":        ("Fecha Registro",      None),
    "Hora de la instrucción \n-recepción-":          ("Hora Registro",       None),
    "Nombre de la persona que gira la instrucción":  ("Nombre",              None),
    "Persona facultada para girar instrucciones":    (None,                  "SI"),
    "Contrato se encuentra vigente":                 (None,                  "SI"),
    "Instrucción registrada como orden":             (None,                  "SI"),
    "Contrato":                                      ("Contrato",            None),
    "Tipo de servicio":                              ("Servicio Contratado", None),
    "Cliente":                                       ("Nombre",              None),
    "Sentido de la operación":                       ("Operación",           None),
    "Emisora":                                       ("Emisora",             None),
    "Serie":                                         ("Serie",               None),
    "Títulos":                                       ("Títulos Ordenados",   None),
    "Precio fijado":                                 ("Precio asignado",     None),
    "Precio a mercado":                              ("Mdo",                 None),
    "Tipo Orden":                                    ("Tipo Orden",          None),
    "Vigencia":                                      ("Vigencia Original",   None),
    "Medio de instrucción":                          ("Medio Instruccion",   None),
    "Clave del promotor que atendió":                ("Operador",            None),
    "Nombre del promotor que atendió":               ("Operador",            "PROMOTOR"),
    "Promotor asignado al contrato":                 ("Contrato",            "CONTRATO_PROMOTOR"),
    "Hora de la captura\n-registro-":                ("Hora Registro",       None),
    "Folio Orden":                                   ("Folio Orden",         None),
    "Comentarios":                                   (None,                  ""),
}

HEADER_ROW = 2
DATA_START  = 3
LAYOUT_PATH = "Layout Bitácora Promotor.xlsx"

# ─────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────
def parse_date(x):
    if pd.isna(x): return None
    if isinstance(x, date): return x
    dt = pd.to_datetime(x, errors="coerce")
    return None if pd.isna(dt) else dt.date()

def parse_time(x):
    if pd.isna(x): return None
    if isinstance(x, time): return x
    dt = pd.to_datetime(str(x), errors="coerce")
    return None if pd.isna(dt) else dt.time()

def get_promotor(row):
    try:
        contrato = int(row["Contrato"])
    except (ValueError, TypeError):
        contrato = None
    return CONTRATO_PROMOTOR.get(contrato, "SIN ASIGNAR")

def build_bitacora(df_p: pd.DataFrame, layout_bytes: bytes) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(layout_bytes))
    ws = wb[wb.sheetnames[0]]

    header_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, c).value
        if isinstance(val, str):
            header_map[val.strip()] = c

    for i, row in enumerate(df_p.to_dict("records")):
        r = DATA_START + i
        for header, (src_col, rule) in LAYOUT_MAP.items():
            h_key = header.strip()
            if h_key not in header_map:
                continue
            c = header_map[h_key]

            if rule == "PROMOTOR":
                val = OP_TO_NAME.get(str(row["Operador"]).strip(), str(row["Operador"]).strip())
            elif rule == "CONTRATO_PROMOTOR":
                val = get_promotor(row)
            else:
                val = row.get(src_col) if src_col else rule

            if header.startswith("Fecha"):
                val = parse_date(val)
            if "Hora" in header:
                val = parse_time(val)

            ws.cell(r, c).value = val

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────
mes = st.text_input("Mes", value="Marzo 2026")

src_file = st.file_uploader("Sube el archivo de Bitácoras (.xlsx)", type=["xlsx"])

if not src_file:
    st.stop()

try:
    with open(LAYOUT_PATH, "rb") as f:
        layout_bytes = f.read()
except FileNotFoundError:
    st.error(f"No se encontró `{LAYOUT_PATH}` en la misma carpeta que este script.")
    st.stop()

src = pd.read_excel(src_file)
src["Operador"] = src["Operador"].astype(str).str.strip()
src["__Promotor__"] = src.apply(get_promotor, axis=1)

# ── Comprobaciones ────────────────────────────────────────────
st.subheader("Comprobaciones")

total       = len(src)
sin_asignar = src[src["__Promotor__"] == "SIN ASIGNAR"]

c1, c2, c3 = st.columns(3)
c1.metric("Total operaciones", total)
c2.metric("Asignadas", total - len(sin_asignar))
c3.metric("Sin asignar", len(sin_asignar),
          delta=f"-{len(sin_asignar)}" if len(sin_asignar) else None,
          delta_color="inverse")

dist = (
    src[src["__Promotor__"] != "SIN ASIGNAR"]
    .groupby("__Promotor__")
    .size()
    .reset_index(name="Operaciones")
    .rename(columns={"__Promotor__": "Promotor"})
    .sort_values("Operaciones", ascending=False)
)
st.dataframe(dist, use_container_width=True, hide_index=True)

if len(sin_asignar) > 0:
    st.warning(f"Contratos no reconocidos: {list(sin_asignar['Contrato'].unique())}")

# ── Descargas ─────────────────────────────────────────────────
st.subheader("Descargar bitácoras")

promotores = [p for p in src["__Promotor__"].unique() if p != "SIN ASIGNAR"]

# ZIP con todos
zip_buf = io.BytesIO()
with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
    for promotor in sorted(promotores):
        df_p = src[src["__Promotor__"] == promotor].copy()
        zf.writestr(f"{promotor} {mes}.xlsx", build_bitacora(df_p, layout_bytes))
zip_buf.seek(0)

st.download_button(
    label=f"⬇️ Descargar todos en ZIP ({len(promotores)} archivos)",
    data=zip_buf,
    file_name=f"Bitacoras_{mes.replace(' ', '_')}.zip",
    mime="application/zip",
    type="primary",
)

# Individuales
st.caption("O descarga por separado:")
for promotor in sorted(promotores):
    df_p = src[src["__Promotor__"] == promotor].copy()
    st.download_button(
        label=f"📄 {promotor}  ({len(df_p)} operaciones)",
        data=build_bitacora(df_p, layout_bytes),
        file_name=f"{promotor} {mes}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=promotor,
    )
